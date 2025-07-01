from django.shortcuts import render
from rest_framework.permissions import IsAuthenticated , IsAdminUser
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from django.contrib.auth import authenticate
from rest_framework import status , permissions
from rest_framework_simplejwt.tokens import RefreshToken
from useracc.models import User
from django.http import JsonResponse
from .serilizers import TheatreSerializer , MovieSerializers , AdminSettingsSerializer
from movies.models import City , Movie
from theatres.models import *
from theatre_owner.models import *
from rest_framework.decorators import api_view
from booking.models import *
from booking.serializers import BookingSerializer
from django.utils import timezone
from collections import OrderedDict
from rest_framework.decorators import permission_classes
from django.db.models import Sum , Count , Q , F , Value as V
from django.db.models.functions import Coalesce
import pandas as pd
from django.db.models.functions import TruncMonth, TruncWeek
import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from django.db.models import ExpressionWrapper, FloatField
from theatre_owner.tasks import send_response_mail
from .models import AdminSettings
 
# Create your views here.

logger = logging.getLogger(__name__)
# customer listed views
class UserListView(APIView):
    permission_classes = [permissions.IsAdminUser]
    
    def get(self , request):
        
        users = User.objects.all().values().order_by('id')
        user_list = list(users)
        return JsonResponse(user_list, safe=False)
    
# user status update view
class UserStatusUpdate(APIView) :
    permission_classes = [permissions.IsAdminUser]

    def post(self , request , pk) :
        try :
            user = User.objects.get(pk=pk)
            user.is_active = not user.is_active
            user.save()
            return Response({'message': 'user status updated successfully'},status=status.HTTP_200_OK)
        except user.DoesNotExist:
            return Response({ 'error' : 'user not found'},status=status.HTTP_404_NOT_FOUND)

class CityTheatreView(APIView) :
    def get(self , request , city_id) :
        try :
            city = City.objects.get(id = city_id)
            
            theatres = Theatre.objects.filter(city = city)
            serialzer = TheatreSerializer(theatres , many=True)
            response_data = {
                "cityname" : city.name,
                "theatres" : serialzer.data
            }
            return Response(response_data , status=status.HTTP_200_OK)
        except City.DoesNotExist:
            return Response({"detail" : 'no theatre found'},status=status.HTTP_404_NOT_FOUND)
        except Exception as e :
            return Response({'detail' : str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# add theatre view
class AddTheatre(APIView) :
    def post(self , request , owner_id ) :

        try :
            city = TheaterOwnerProfile.objects.get(id = owner_id)

        except TheaterOwnerProfile.DoesNotExist :
            return Response({ "detail" : "city not found" },status=status.HTTP_404_NOT_FOUND)

        name = request.data.get('name')
        address = request.data.get('address')

        if not name or not address :
            return Response({"detail" : "name and address are required"},status=status.HTTP_400_BAD_REQUEST)
        

        try :

            theatre = Theatre.objects.create(

                name = name,
                city = city ,
                address = address
            )
            return Response(
                {'detail' : f"theatre {theatre.name} created successfully"},
                status=status.HTTP_201_CREATED
            )

        except Exception as e :
            return Response(
                {"detail" : f"An erorr occurred : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class DeleteTheatre(APIView) :
    def delete(self , request , id):
    
        try :
            theatre = Theatre.objects.get(id = id)
        except Theatre.DoesNotExist :
            return Response({"error": "Theatre not found"}, status=status.HTTP_404_NOT_FOUND)
        
        theatre.delete()
        return Response({"message": "Theatre deleted successfully"}, status=status.HTTP_200_OK)

# Movies View ( CRUD )
class ListMovies(APIView) :
    def get(self , reqeust) :
        queryset = Movie.objects.all()

        movie_list = []
        for movie in queryset :

            movie_dict = {

                'id': movie.id,
                'title': movie.title,
                'language': movie.language,
                'duration': movie.duration,
                'release_date': movie.release_date,
                'description': movie.description,
                'genre': movie.genre,
                'poster' : reqeust.build_absolute_uri(movie.poster.url) if movie.poster else None
            }

            movie_list.append(movie_dict)
        return JsonResponse(movie_list , safe=False)

        
# create movie view 
class CreateMovieView(APIView) :
    def post(self , request) :
        data = request.data
        serializer = MovieSerializers(data= data)
        if serializer.is_valid() :
            serializer.save()
            return Response(
                {"message" : "Movie created successfully" , "data" : serializer.data},
                status=status.HTTP_201_CREATED,    
            )
        print(serializer.errors)
        return Response(
            serializer.errors ,status=status.HTTP_400_BAD_REQUEST
        )
# update movie view
class update_movie(APIView):     
    def get(self ,request , movie_id):   
        try:
            movie = Movie.objects.get(id=movie_id)
        except Movie.DoesNotExist:
            return Response({"error": "Movie not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = MovieSerializers(movie)
        return Response(serializer.data)
    
    def put(self , request , movie_id):
        try:
            movie = Movie.objects.get(id=movie_id)
        except Movie.DoesNotExist:
            return Response({"error": "Movie not found"}, status=status.HTTP_404_NOT_FOUND)
       
        serializer = MovieSerializers(movie, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Movie updated successfully", "movie": serializer.data})
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
# delete movie view
class DeleteMovies(APIView) :
    def delete(self , requst , id):
        try :
            movie = Movie.objects.get(id=id)
            
        except Movie.DoesNotExist:
            return Response({'error' : 'movie not found'} , status=status.HTTP_404_NOT_FOUND)
        movie.delete()
        return Response({'message' : f'{movie.title} deleted successfully'} , status=status.HTTP_200_OK)
    
# show pending theatres view including screens
class ShowTheatreRequest(APIView):
    def get(self , request) :   
        try :
            
            theatres = Theatre.objects.filter(screens__is_approved=False )
        except Theatre.DoesNotExist:
            return Response({'message' : 'pending theatres not found'},status=status.HTTP_404_NOT_FOUND)
        
        serializer = TheatreSerializer(theatres , many=True)
        print(serializer.data)
        return Response(serializer.data , status=status.HTTP_200_OK)
    
    def post(self , request ):
        try :   
            theatre_id = request.data.get('theatre_id')
            theatre_det = Theatre.objects.get(id = theatre_id)
            action = request.data.get('action')
            if action == 'confirmed':
                if not theatre_det.has_screen():
                    return Response(
                        {'error':'Theatre not have at least one screen details'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                theatre_det.is_confirmed = True
                theatre_det.save()
                
                screen_data = Screen.objects.filter(theatre=theatre_det,is_approved=False)
                if screen_data :
                    for screen in screen_data :
                        screen.is_approved = True
                        screen.save()
                        
                subject = f"{theatre_det.name} screen was approved by admin"
                msg = f"thank you for being part of cinemagic" 
                send_response_mail(
                    theatre_det.owner.user.email,
                    subject,
                    msg
                )
                return Response(
                    {'message' : 'theatre verified successfully'},
                    status=status.HTTP_200_OK
                )
                
        except Theatre.DoesNotExist:
            return Response(
                {'error':'theatre not found'},
                status=status.HTTP_404_NOT_FOUND
            )
# show verified theatres view
@api_view(['get'])
def Verified_Theatres(request):
    try :
        theatres = Theatre.objects.filter(is_confirmed=True).order_by('-created_at')
    except Theatre.DoesNotExist:
        return Response({'error' : 'no theatres found'},status=status.HTTP_404_NOT_FOUND)
            
    serializers = TheatreSerializer(theatres , many=True)
    return Response(serializers.data , status=status.HTTP_200_OK)

# check theatre verfication view
class verify_screen(APIView):
    def post(self , request,screen_id):
        data = request.data
        owner = data.get('owner_id')
        owner = TheaterOwnerProfile.objects.get(id=owner)
        try :
            screen = Screen.objects.get(id=screen_id)
            screen.is_approved = True
            screen.save()
            subject = f"{screen.theatre.name} screen {screen.screen_number} was approved by admin"
            email_msg = f'screen updated by cineMagic admins \n\n wish you a good start🎉'

            send_response_mail(
                owner.user.email,
                subject , 
                email_msg
                
            )
            return Response({'message' : 'approved successfully'},status=status.HTTP_200_OK)
        
        except Screen.DoesNotExist:
            return Response({'error' : 'not found screen'}, status=status.HTTP_404_NOT_FOUND)
            
            
    def delete(self , request , screen_id):
        try :
                    
            screen = Screen.objects.get(id=screen_id)
        except Screen.DoesNotExist:
            return Response({'error' : 'screen not found'},status=status.HTTP_404_NOT_FOUND)
        screen_number = screen.screen_number
        theatre_name = screen.theatre.name
        screen.delete()
            
        return Response({'message' : f'Screen : {screen_number} from {theatre_name} deleted successfully  '},status=status.HTTP_200_OK)
    
# delete show view
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def Cancel_Show(request , show_id) :
    show = ShowTime.objects.get(id=show_id)
    screen_number = show.screen.screen_number
    movie = show.movie.title
    show.delete()
    return Response({'message' : f'{movie} on screen {screen_number} was successfully cancelled'})

# pending cancelled shows view
@permission_classes([IsAdminUser])
class PendingCancelledShows(APIView):
    def get(self , request) :
        try : 
            bookings = Booking.objects.all().exclude(status='not-applicable').filter(status='cancelled').order_by('-booking_time')
            
        except Booking.DoesNotExist:
            return Response({'error' : 'not booking found'},status=status.HTTP_404_NOT_FOUND)
        
        
        serializer = BookingSerializer(bookings , many=True)
        return Response(serializer.data , status=status.HTTP_200_OK)
    
# dashboard status information (admin view)

class dashboard_stats(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self , request) :
        today = timezone.now()
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date :
            start_date = today - timedelta(days=30)
        else : 
            start_date = datetime.fromisoformat(start_date)
        if not end_date :
            end_date = today
        else :
            end_date = datetime.fromisoformat(end_date)
            
        booking_queryset = Booking.objects.filter(
            booking_time__range=[start_date , end_date],
            status='confirmed'
        )
        
        total_revenue = booking_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        print(total_revenue)
        total_tickets = BookingSeat.objects.filter(
            booking__in = booking_queryset,
            status='booked'
        ).count()
        
        active_users = User.objects.filter(booking__in=booking_queryset).distinct().count()
        
        active_theatres = Theatre.objects.filter(
            screens__showtimes__booking__in=booking_queryset,
            is_confirmed=True
        ).distinct().count()
        
        # calculate percentage changes (compared to previoous period)

        previous_stats = start_date - (end_date - start_date)
        previous_booking_queryset = Booking.objects.filter(
            booking_time__range=[previous_stats , start_date],
            status='confirmed'
        )
        
        previous_revenue = previous_booking_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        
        revenue_chenge = ((total_revenue - previous_revenue) / previous_revenue) * 100 if previous_revenue else 0
        ticket_change = ((total_tickets - previous_booking_queryset.count()) / previous_booking_queryset.count()) * 100 if previous_booking_queryset.count() else 0
        print(ticket_change , 'changes from last month')
        
        return Response({
            'total_tickets' : total_tickets,
            'active_users' : active_users,
            'active_theatres' : active_theatres,
            'revenue_change' : revenue_chenge,
            'ticket_change' : ticket_change,
            'period' : {
                'start_date' : start_date,
                'end_date' : end_date
            }
        },status=status.HTTP_200_OK)
        
        
        
class revenue_chart_data(APIView):
    permission_classes = [permissions.AllowAny]
    
    def get(self , request):
        period = request.GET.get('period' , 'month')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        
        logger.info(f"Received period: {period}, start_date: {start_date}, end_date: {end_date}")
        
        
        if not start_date and not end_date:
            print('no start date and end date')
            if period == 'week':
                start_date = timezone.now() - timedelta(days=7)
                end_date = timezone.now()
                
            if period == 'month':
                start_date = timezone.now() - timedelta(days=30)
                end_date = timezone.now()
            else :
                start_date = timezone.now() - timedelta(days=365)
                end_date = timezone.now()
                
                
        else:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
            
        
        
        if period == 'week':
            trunc_func = TruncWeek
            date_format = '%Y-W%W'
        elif period == 'year':
            trunc_func = TruncMonth  
            date_format = '%Y-%m'
        else:  
            trunc_func = TruncMonth
            date_format = '%b'

        revenue_data = Booking.objects.filter(
            booking_time__range=[start_date, end_date],
            status='confirmed'
        ).annotate(
            period=trunc_func('booking_time')
        ).values('period').annotate(
            total_revenue=Sum('amount'),
            admin_revenue = ExpressionWrapper(Sum('amount') * 0.10, output_field=FloatField()),
            theatre_revenue=ExpressionWrapper(Sum('amount') * 0.90, output_field=FloatField()),
            tickets=Count('bookingseats', filter=Q(bookingseats__status='booked')),
            theatres=Count('show__screen__theatre', distinct=True)
        ).order_by('period')
        
        logger.info(f"Revenue data: {revenue_data}")
        chart_data = []
        for item in revenue_data:
            if period == 'month':
                label = item['period'].strftime('%b')
            elif period == 'week':
                label = f"Week {item['period'].strftime('%W')}"
            else: 
                label = item['period'].strftime('%Y-%m')
                
            chart_data.append({
                'period': label,
                'total_revenue': float(item['total_revenue'] or 0),
                'admin_revenue': float(item['admin_revenue'] or 0),
                'theatre_revenue': float(item['theatre_revenue'] or 0),
                'tickets': item['tickets'],
                'theatres': item['theatres'],
            })
            
        return Response(chart_data , status=status.HTTP_200_OK)
    
    
class ticket_trend_data(APIView):
    def get(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date:
            start_date = timezone.now() - timedelta(days=365)
        else:
            start_date = datetime.fromisoformat(start_date)
        
        if not end_date:
            end_date = timezone.now()
        else:
            end_date = datetime.fromisoformat(end_date)

        # Query aggregated ticket data grouped by month
        ticket_data = BookingSeat.objects.filter(
            booking__booking_time__range=[start_date, end_date],
            booking__status='confirmed',
            status='booked'
        ).annotate(
            month=TruncMonth('booking__booking_time')
        ).values('month').annotate(
            tickets=Count('id')
        ).order_by('month')

        ticket_dict = {item['month'].strftime('%b'): item['tickets'] for item in ticket_data}

        result = []
        month_pointer = start_date.replace(day=1)
        while month_pointer <= end_date:
            month_abbr = month_pointer.strftime('%b')
            result.append({
                'month': month_abbr,
                'tickets': ticket_dict.get(month_abbr, 0)
            })
            if month_pointer.month == 12:
                month_pointer = month_pointer.replace(year=month_pointer.year + 1, month=1)
            else:
                month_pointer = month_pointer.replace(month=month_pointer.month + 1)

        return Response(result, status=status.HTTP_200_OK)
    
    
class RecentSale(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self , request):
        print(request.data)
        limit = int(request.data.get('limit' , 5))
        try:
            recent_bookings = Booking.objects.filter(
                status='confirmed'
            ).select_related(
                'user', 'show__movie', 'show__screen__theatre'
            ).order_by('-booking_time')[:limit]
            
            logger.info(recent_bookings)
            
            sales_data = []
            for booking in recent_bookings:
                time_diff = timezone.now() - booking.booking_time
                if time_diff.days > 0:
                    time_ago = f"{time_diff.days}d ago"
                elif time_diff.seconds > 3600:
                    hours = time_diff.seconds // 3600
                    time_ago = f"{hours}h ago"
                else:
                    minutes = time_diff.seconds // 60
                    time_ago = f"{minutes}m ago"
                
                sales_data.append({
                    'id': booking.booking_id,
                    'customer': booking.customer_name,
                    'email': booking.customer_email,
                    'amount': float(booking.amount),
                    'movie': booking.show.movie.title if booking.show and booking.show.movie else 'N/A',
                    'theatre': booking.show.screen.theatre.name if booking.show and booking.show.screen else 'N/A',
                    'time': time_ago,
                    'tickets_count': booking.bookingseats.filter(status='booked').count()
                })
            
            return Response(sales_data , status=status.HTTP_200_OK)
        
        except Exception as e:
            print(f"Error fetching recent sales: {str(e)}") 
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
# Download xl sheet Reports view 
class ExportTheatreReport(APIView):
    def get(self , request):
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        theatre_id = request.GET.get('theatre_id')
        report_type = request.GET.get('report_type' , 'all')
        
        logger.info(f'data inside export Theatre report ' , request.data)
        
        
        if not start_date :
            start_date = timezone.now() - timedelta(days=30)
        else :
            start_date = datetime.fromisoformat(start_date)
            
        if not end_date :
            end_date = timezone.now()
        else : 
            end_date = datetime.fromisoformat(end_date)
            
            
        bookings_queryset = Booking.objects.filter(
            booking_time__range=[start_date , end_date],
            status='confirmed',
        )

        
        if theatre_id:
            try:
                theatre = Theatre.objects.get(id=theatre_id)
                bookings_queryset = bookings_queryset.filter(show__screen__theatre=theatre)
            except Theatre.DoesNotExist:
                return Response({'error': 'Theatre not found'}, status=status.HTTP_404_NOT_FOUND)
            
            
        wb = openpyxl.Workbook()
    
        wb.remove(wb.active)
        
        if report_type == 'all' or report_type == 'seat_analysis':
            create_seat_category_sheet(wb, bookings_queryset, start_date, end_date)

        if report_type == 'all' or report_type == 'detailed_performance':
            create_detailed_performance_sheet(wb, bookings_queryset, start_date, end_date)

        if report_type == 'all' or report_type == 'monthly_breakdown':
            create_monthly_breakdown_sheet(wb, bookings_queryset, start_date, end_date)

        if report_type == 'all' or report_type == 'theatre_summary':
            create_theatre_summary_sheet(wb, bookings_queryset, start_date, end_date)

        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'theatre_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response 
    
def create_theatre_summary_sheet(wb, bookings_query, start_date, end_date):
    """Create theatre summary sheet"""
    ws = wb.create_sheet("Theatre Summary")
    
    theatre_data = []
    theatres = Theatre.objects.filter(
        screens__showtimes__booking__in=bookings_query,
        is_confirmed=True
    ).distinct()
    
    for theatre in theatres:
        theatre_bookings = bookings_query.filter(
            show__screen__theatre=theatre
        )
        
        total_revenue = theatre_bookings.aggregate(Sum('amount'))['amount__sum'] or 0
        total_tickets = BookingSeat.objects.filter(
            booking__in=theatre_bookings,
            status='booked'
        ).count()
        
        total_shows = theatre_bookings.values('show').distinct().count()
        avg_ticket_price = (total_revenue / total_tickets) if total_tickets > 0 else 0
        
        theatre_data.append({
            'Theatre Name': theatre.name,
            'City': theatre.city.name,
            'Address': theatre.address,
            'Total Revenue': float(total_revenue),
            'Tickets Sold': total_tickets,
            'Total Shows': total_shows,
            'Average Ticket Price': float(avg_ticket_price),
            'Revenue per Show': float(total_revenue / total_shows) if total_shows > 0 else 0
        })
    
    df = pd.DataFrame(theatre_data)
    
    ws.append(['THEATRE PERFORMANCE SUMMARY'])
    ws.append([f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'])
    
    if theatre_data is None :
        ws.append(['No theatre tickets Sold selected Range'])
    ws.append([])
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    style_worksheet(ws, df.shape[0] + 4)
    


def create_seat_category_sheet(wb, bookings_query, start_date, end_date):
    """Create seat category analysis sheet"""
    ws = wb.create_sheet("Seat Category Analysis")
    
    seat_data = []
    theatres = Theatre.objects.filter(
        screens__showtimes__booking__in=bookings_query,
        is_confirmed=True
    ).distinct()
    
    for theatre in theatres:
        theatre_bookings = bookings_query.filter(
            show__screen__theatre=theatre
        )
        
        # Get seat categories for this theatre
        seat_categories = BookingSeat.objects.filter(
            booking__in=theatre_bookings,
            status='booked'
        ).annotate(
            category_name=Coalesce(F('seat__category__name'), V('Standard'))

        ).values(
            
            'category_name'    
        ).annotate(
            
            count=Count('id'),
            revenue=Sum('price')
        )
        
        for category in seat_categories:
            seat_data.append({
                'Theatre': theatre.name,
                'City': theatre.city.name,
                'Seat Category': category['category_name'] or 'Standard',
                'Tickets Sold': category['count'],
                'Revenue': float(category['revenue'] or 0),
                'Average Price': float(category['revenue'] / category['count']) if category['count'] > 0 else 0
            })
    
    df = pd.DataFrame(seat_data)
    
    if not df.empty:
        category_summary = df.groupby('Seat Category').agg({
            'Tickets Sold': 'sum',
            'Revenue': 'sum'
        }).reset_index()
        category_summary['Average Price'] = category_summary['Revenue'] / category_summary['Tickets Sold']
    
    ws.append(['SEAT CATEGORY ANALYSIS'])
    ws.append([f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'])
    ws.append([])
    
    ws.append(['DETAILED BREAKDOWN BY THEATRE'])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    ws.append([])
    ws.append(['CATEGORY SUMMARY'])
    if not df.empty:
        for row in dataframe_to_rows(category_summary, index=False, header=True):
            ws.append(row)
    
    style_worksheet(ws, len(df) + len(category_summary) + 8 if not df.empty else 8)

def create_detailed_performance_sheet(wb, bookings_query, start_date, end_date):
    """Create detailed performance sheet with movie-wise data"""
    ws = wb.create_sheet("Detailed Performance")
    
    detailed_data = []
    
    bookings = bookings_query.select_related(
        'show__movie', 'show__screen__theatre', 'show__screen__theatre__city'
    ).prefetch_related('bookingseats')
    
    for booking in bookings:
        if booking.show and booking.show.movie and booking.show.screen:
            seats_booked = booking.bookingseats.filter(status='booked')
            
            detailed_data.append({
                'Booking ID': booking.booking_id,
                'Theatre': booking.show.screen.theatre.name,
                'City': booking.show.screen.theatre.city.name,
                'Movie': booking.show.movie.title,
                'Screen': f"Screen {booking.show.screen.screen_number}",
                'Customer Name': booking.customer_name,
                'Customer Email': booking.customer_email,
                'Booking Date': booking.booking_time.strftime('%Y-%m-%d'),
                'Show Date': booking.show.show_date.strftime('%Y-%m-%d') if booking.show.show_date else 'N/A',
                'Amount': float(booking.amount),
                'Seats Count': seats_booked.count(),
                'Status': booking.status
            })
    
    df = pd.DataFrame(detailed_data)
    
    # Add headers
    ws.append(['DETAILED BOOKING PERFORMANCE'])
    ws.append([f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'])
    ws.append([])
    
    # Add data
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    # Style the sheet
    style_worksheet(ws, len(df) + 4)

def create_monthly_breakdown_sheet(wb, bookings_query, start_date, end_date):
    """Create monthly breakdown sheet"""
    ws = wb.create_sheet("Monthly Breakdown")
    
    logger.info(f'inside monthly break down function here' , ws.title)
    
    monthly_data = []
    theatres = Theatre.objects.filter(
        screens__showtimes__booking__in=bookings_query,
        is_confirmed=True
    ).distinct()
    
    for theatre in theatres:
        theatre_bookings = bookings_query.filter(
            show__screen__theatre=theatre
        )
        
        monthly_stats = theatre_bookings.extra(
            select={'month': "DATE_TRUNC('month', booking_time)"}
        ).values('month').annotate(
            revenue=Sum('amount'),
            tickets=Count('bookingseats', filter=Q(bookingseats__status='booked')),
            bookings_count=Count('id')
        ).order_by('month')
        
        for stat in monthly_stats:
            monthly_data.append({
                'Theatre': theatre.name,
                'City': theatre.city.name,
                'Month': stat['month'].strftime('%Y-%m'),
                'Revenue': float(stat['revenue'] or 0),
                'Tickets Sold': stat['tickets'],
                'Total Bookings': stat['bookings_count'],
                'Avg Booking Value': float(stat['revenue'] / stat['bookings_count']) if stat['bookings_count'] > 0 else 0
            })
    
    df = pd.DataFrame(monthly_data)
    ws.append(['MONTHLY BREAKDOWN'])
    ws.append([f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'])
    ws.append([])
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    style_worksheet(ws, len(df) + 4)
    
def create_seat_category_sheet(wb, bookings_query, start_date, end_date):
    """Create seat category analysis sheet"""
    ws = wb.create_sheet("Seat Category Analysis")
    print(bookings_query)
    seat_data = []
    theatres = Theatre.objects.filter(
        screens__showtimes__booking__in=bookings_query,
        is_confirmed=True
    ).distinct()
    
    for theatre in theatres:
        theatre_bookings = bookings_query.filter(
            show__screen__theatre=theatre
        )
        
        # Get seat categories for this theatre
        seat_categories = BookingSeat.objects.filter(
            booking__in=theatre_bookings,
            status='booked'
        ).values(
            'seat__category'  
        ).annotate(
            count=Count('id'),
            revenue=Sum('price')
        )
        
        for category in seat_categories:
            seat_data.append({
                'Theatre': theatre.name,
                'City': theatre.city.name,
                'Seat Category': category['seat__category'] or 'Standard',
                'Tickets Sold': category['count'],
                'Revenue': float(category['revenue'] or 0),
                'Average Price': float(category['revenue'] / category['count']) if category['count'] > 0 else 0
            })
    
    df = pd.DataFrame(seat_data)
    
    if not df.empty:
        category_summary = df.groupby('Seat Category').agg({
            'Tickets Sold': 'sum',
            'Revenue': 'sum'
        }).reset_index()
        category_summary['Average Price'] = category_summary['Revenue'] / category_summary['Tickets Sold']
    
    ws.append(['SEAT CATEGORY ANALYSIS'])
    ws.append([f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'])
    ws.append([])
    
    ws.append(['DETAILED BREAKDOWN BY THEATRE'])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    ws.append([])
    ws.append(['CATEGORY SUMMARY'])
    if not df.empty:
        for row in dataframe_to_rows(category_summary, index=False, header=True):
            ws.append(row)
    
    style_worksheet(ws, len(df) + len(category_summary) + 8 if not df.empty else 8)

    


def style_worksheet(ws, data_rows):
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center')
    
    header_row = None
    for idx, row in enumerate(ws.iter_rows(), 1):
        if any(cell.value for cell in row if isinstance(cell.value, str) and 
               cell.value in ['Theatre Name', 'Theatre', 'Booking ID']):
            header_row = idx
            break
    
    if header_row:
        for cell in ws[header_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        for row in ws.iter_rows(min_row=header_row, max_row=data_rows):
            for cell in row:
                cell.border = thin_border
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50) 
        ws.column_dimensions[column_letter].width = adjusted_width
        
        
api_view(['GET'])
def GetTheatres(request):
    bookings = Booking.objects.all().select_related('show__screen__theatre')
    data = []
    seen_theatre_id = set()
    for booking in bookings :
        screen = booking.show.screen
        theatre = screen.theatre
        if theatre.id not in seen_theatre_id:
            seen_theatre_id.add(theatre.id)
            data.append({
                'id' : theatre.id,
                'name' : theatre.name,
                'city' : theatre.city.name
            })
    
    
    return JsonResponse(data , safe=False, status=200)

    
class AdminSettingsView(APIView):
    def get(self , request):
        settings_ob , created = AdminSettings.objects.get_or_create(id=1)
        serializer = AdminSettingsSerializer(settings_ob)
        return Response(serializer.data , status=status.HTTP_200_OK)
    
    
    def post(self, request):
        settings_obj, created = AdminSettings.objects.get_or_create(id=1)
        serializer = AdminSettingsSerializer(settings_obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Settings updated successfully!', 'data': serializer.data},status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ChangePassword(APIView):
    def post(self , request ):
        try:
            user = request.user
            username = user.username
            old_password = request.data.get('old_password')
            new_password = request.data.get('new_password')
            user = authenticate(username=username , password=old_password)
            
            if user and user.is_superuser:
                user.set_password(new_password)
                user.save()
                return Response({'message':'password changed'},status=status.HTTP_200_OK)
            return Response({'message':'invalid password or not a superuser'})
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    